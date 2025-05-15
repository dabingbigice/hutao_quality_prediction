import pandas as pd
from openpyxl import load_workbook


def append_to_excel(file_path, new_data):
    """
    向Excel文件追加新数据
    参数：
    file_path: 现有Excel文件路径
    new_data: 要追加的数据列表，格式示例：
        [
            [1001, 85.3, 25.3, 22.0],  # area_num, perimeter, a, b
            [1002, 92.1, 27.1, 25.1]
        ]
    """
    # 转换数据为DataFrame
    df_new = pd.DataFrame(new_data, columns=['area_num', 'perimeter', 'a', 'b', 'a/b', 'area_num/perimeter', 'g','error'])

    # 自动生成计算列
    # df_new['a/b'] = df_new['a'] / df_new['b']
    # df_new['area_num/perimeter'] = df_new['area_num'] / df_new['perimeter']
    # df_new['g'] = new_data[0][4]
    # 使用openpyxl引擎追加数据
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        # 找到最后一行
        book = load_workbook(file_path)
        sheet = book.active
        start_row = sheet.max_row

        # 追加数据（跳过表头）
        df_new.to_excel(writer, startrow=start_row, index=False, header=False)
        print(f'excel保存成功={df_new}')


# 使用示例 ---------------------------------------------------
if __name__ == "__main__":
    # 新数据示例（可修改）
    new_data = [
        [1003, 78.4, 23.8, 19.5],  # 数据对应area_num, perimeter, a, b,a/b,a/p
        [1004, 81.2, 24.5, 20.1]
    ]

    # 文件路径（修改为实际路径）
    excel_path = "核桃仁表型信息.xlsx"

    # 执行追加
    append_to_excel(excel_path, new_data)
    print(f"数据已成功追加到 {excel_path}")
